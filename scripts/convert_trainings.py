from __future__ import annotations
import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "(10ºMesociclo) MIX OF LOADS II    11-09-23.xlsx"
OUTPUT = ROOT / "public" / "data" / "trainings.json"

MICROCYCLE_COLUMN_INDEXES = [2, 4, 6, 8, 10, 12]


def normalize(value: Optional[Any]) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def extract_microcycle_labels(rows: List[List[Optional[str]]]) -> List[str]:
    labels: List[str] = []
    if len(rows) > 4:
        header_row = rows[4]
        for idx in MICROCYCLE_COLUMN_INDEXES:
            if idx < len(header_row):
                cell = normalize(header_row[idx])
                if cell:
                    labels.append(cell)
    return labels


def parse_warmups(rows: List[List[Optional[str]]]) -> List[Dict[str, Optional[str]]]:
    warmups: List[Dict[str, Optional[str]]] = []
    inside = False
    for row in rows:
        text = normalize(row[1]) if len(row) > 1 else None
        if text == "EN EL CALENTAMIENTO:":
            inside = True
            continue
        if inside:
            if not text:
                break
            if text and text.startswith("SERIES &"):
                break
            link = None
            if len(row) > 4:
                link = normalize(row[4])
            warmups.append({"description": text, "resource": link})
    return warmups


def is_exercise_header(row: List[Optional[str]]) -> bool:
    if len(row) < 3:
        return False
    name = normalize(row[1])
    marker = normalize(row[2])
    if not name or not marker:
        return False
    return "X" in marker.upper() and not name.lower().startswith("serie")


def is_series_row(label: Optional[str]) -> bool:
    if not label:
        return False
    norm = label.strip().lower()
    return norm.startswith("1ª serie") or norm.startswith("2ª serie") or norm.startswith("3ª serie") or norm.startswith("4ª serie") or norm.startswith("5ª serie")


def parse_exercises(rows: List[List[Optional[str]]]) -> List[Dict[str, Any]]:
    exercises: List[Dict[str, Any]] = []
    i = 0
    # advance to first exercise section
    while i < len(rows):
        text = normalize(rows[i][1]) if len(rows[i]) > 1 else None
        if text == "SERIES & REPETICIONES TARGET":
            i += 1
            break
        i += 1

    while i < len(rows):
        row = rows[i]
        if is_exercise_header(row):
            name = normalize(row[1])
            header = [normalize(cell) for cell in row[:14]]
            exercise = {
                "name": name,
                "header": header,
                "notes": [],
                "series": [],
                "rest": None,
            }
            i += 1
            while i < len(rows):
                current = rows[i]
                current_name = normalize(current[1]) if len(current) > 1 else None
                if current_name == "SERIES & REPETICIONES TARGET":
                    i += 1
                    continue
                if is_exercise_header(current):
                    break
                row_values = [normalize(cell) for cell in current[:14]]
                if current_name == "Tiempo de pausa entre series":
                    exercise["rest"] = row_values
                elif is_series_row(current_name):
                    exercise["series"].append(row_values)
                else:
                    if any(value for value in row_values):
                        exercise["notes"].append(row_values)
                i += 1
            exercises.append(exercise)
            continue
        i += 1
    return exercises


def parse_sheet(sheet_name: str) -> Dict[str, Any]:
    wb = load_workbook(SOURCE, data_only=True)
    ws = wb[sheet_name]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    str_rows = [[normalize(cell) for cell in row] for row in rows]
    phase = str_rows[1][1] if len(str_rows) > 1 and len(str_rows[1]) > 1 else None
    title = str_rows[3][1] if len(str_rows) > 3 and len(str_rows[3]) > 1 else None
    microcycles = extract_microcycle_labels(str_rows)
    warmups = parse_warmups(str_rows)
    exercises = parse_exercises(str_rows)
    return {
        "sheet": sheet_name,
        "phase": phase,
        "title": title,
        "microcycles": microcycles,
        "warmups": warmups,
        "exercises": exercises,
    }


def main() -> None:
    if not SOURCE.exists():
        raise SystemExit(f"Source workbook not found: {SOURCE}")
    wb = load_workbook(SOURCE, data_only=True, read_only=True)
    payload = {sheet: parse_sheet(sheet) for sheet in wb.sheetnames}
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT.open("w", encoding="utf-8") as fp:
        json.dump(payload, fp, ensure_ascii=False, indent=2)
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
